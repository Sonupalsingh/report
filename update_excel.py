import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os

# File paths
txt_file_path = r"C:\Users\Indo\Desktop\auto\system_details.txt"
excel_file_path = r"C:\Users\Indo\Desktop\auto\auto.xlsx"

# Read the system details from the text file
with open(txt_file_path, 'r') as file:
    data = file.read()

# Extract details for each server
hosts = re.findall(r"Host:\s*(\S+)", data)
hostnames = re.findall(r"Hostname:\s*(\S+)", data)
ip_addresses = re.findall(r"IP Address:\s*(\S+)", data)
date_times = re.findall(r"Date & Time:\s*([0-9\-]+\s+[0-9:]+)", data)
memory_usages = re.findall(r"Memory Usage:\s*.*?Usage:\s*([0-9\.]+)%", data, re.DOTALL)
cpu_usages = re.findall(r"CPU Usage:\s*Usage:\s*([0-9\.]+)%", data)

# Extract filesystem details block
filesystem_blocks = re.findall(r"Filesystem Utilization:\s*([\s\S]+?)(?=Host:|$)", data)

# Extract and process filesystem details
def extract_filesystem_details(filesystem_data):
    pattern = r"(\S+)\s+\S+\s+\S+\s+(\d+)%\s+(\S+)"
    results = re.findall(pattern, filesystem_data)
    
    formatted_details = []
    for fs, usage, mount in results:
        if int(usage) > 10:
            formatted_details.append(f"{fs} {usage}% {mount}")

    return "OK" if not formatted_details else "\n".join(formatted_details)

# Get the current date and time
current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Check if the Excel file exists, and create it if not
if not os.path.exists(excel_file_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "System Details"

    # Add the top headers (Date & Time)
    sheet.merge_cells("A1:H1")
    sheet["A1"] = current_datetime
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Add **corrected** table headers starting in row 3
    headers = ["S.No", "Host", "Hostname", "IP Address", "Filesystem Details",
               "Memory Usage (%)", "CPU Usage (%)", "Report Date & Time"]
    
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col_num, value=header)

    wb.save(excel_file_path)

# Load the Excel workbook
wb = load_workbook(excel_file_path)
sheet = wb.active

# Fix for 'Set changed size during iteration' error
merged_cells_copy = list(sheet.merged_cells.ranges)

# Unmerge all merged cells
for merged_cell_range in merged_cells_copy:
    sheet.unmerge_cells(str(merged_cell_range))

# Remove all data except headers
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

# Find the last used row to continue numbering
start_row = 4

for i in range(len(hosts)):
    row = start_row + i

    sheet[f'A{row}'] = i + 1
    sheet[f'B{row}'] = hosts[i] if i < len(hosts) else ''
    sheet[f'C{row}'] = hostnames[i] if i < len(hostnames) else ''
    sheet[f'D{row}'] = ip_addresses[i] if i < len(ip_addresses) else ''
    sheet[f'E{row}'] = extract_filesystem_details(filesystem_blocks[i].strip()) if i < len(filesystem_blocks) else ''
    sheet[f'F{row}'] = memory_usages[i] if i < len(memory_usages) else ''
    sheet[f'G{row}'] = cpu_usages[i] if i < len(cpu_usages) else ''
    sheet[f'H{row}'] = date_times[i] if i < len(date_times) else ''

# Save the updated workbook
wb.save(excel_file_path)

print(f"? Excel file updated successfully with properly aligned headers!")
